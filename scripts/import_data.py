import openpyxl
import requests
import json
import os

EXCEL_FILE = "../影刀社区帖子数据step03.xlsx"
API_ENDPOINT = "https://backend.appmiaoda.com/projects/supabase335579975837069312/functions/v1/receive-posts"
API_KEY = "yd-rpa-2026-community-dashboard"

def read_excel(file_path):
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    
    headers = [cell.value for cell in ws[1]]
    print("Excel 表头:", headers)
    
    posts = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        post = {}
        for i, header in enumerate(headers):
            if header and row[i] is not None:
                value = row[i]
                if isinstance(value, (int, float)):
                    value = int(value)
                post[header.strip()] = value
        
        if post:
            posts.append(post)
    
    print(f"共读取 {len(posts)} 条帖子数据")
    return posts

def transform_posts(raw_posts):
    tag_map = {
        '网页交互': 'T1',
        '指令报错': 'T2',
        '数据处理': 'T3',
        '第三方对接': 'T4',
        '功能咨询': 'T5',
        'T1': 'T1',
        'T2': 'T2',
        'T3': 'T3',
        'T4': 'T4',
        'T5': 'T5',
    }
    
    sentiment_map = {
        '积极': 'positive',
        '中性': 'neutral',
        '消极': 'negative',
        'positive': 'positive',
        'neutral': 'neutral',
        'negative': 'negative',
    }
    
    posts = []
    for raw in raw_posts:
        post = {
            'title': str(raw.get('帖子标题', '')),
            'content': str(raw.get('帖子内容', '')),
            'author': str(raw.get('帖子作者', '匿名用户')),
            'views': int(raw.get('帖子浏览量', 0)),
            'tag': tag_map.get(str(raw.get('帖子标签分类', 'T5')), 'T5'),
            'sentiment': sentiment_map.get(str(raw.get('帖子情感倾向', 'neutral')), 'neutral'),
            'post_date': str(raw.get('发帖日期', '')),
        }
        posts.append(post)
    
    return posts

def push_to_api(posts):
    headers = {
        'Content-Type': 'application/json',
        'x-api-key': API_KEY,
    }
    
    batch_size = 50
    total_inserted = 0
    
    for i in range(0, len(posts), batch_size):
        batch = posts[i:i+batch_size]
        payload = {'posts': batch}
        
        try:
            response = requests.post(API_ENDPOINT, headers=headers, json=payload, timeout=30)
            result = response.json()
            
            if result.get('success'):
                inserted = result.get('inserted', 0)
                total_inserted += inserted
                print(f"批次 {i//batch_size + 1}: 成功插入 {inserted} 条")
                
                if result.get('failed'):
                    print(f"  失败 {result.get('failed')} 条")
            else:
                print(f"批次 {i//batch_size + 1} 失败: {result.get('error', '未知错误')}")
        
        except Exception as e:
            print(f"批次 {i//batch_size + 1} 请求异常: {str(e)}")
    
    print(f"\n数据导入完成，总计插入 {total_inserted} 条")
    return total_inserted

def main():
    print("=" * 60)
    print("影刀社区帖子数据导入工具")
    print("=" * 60)
    
    if not os.path.exists(EXCEL_FILE):
        print(f"错误: 未找到 Excel 文件 {EXCEL_FILE}")
        return
    
    raw_posts = read_excel(EXCEL_FILE)
    
    if not raw_posts:
        print("错误: 未读取到任何数据")
        return
    
    posts = transform_posts(raw_posts)
    
    print("\n数据转换完成，准备推送...")
    print(f"推送地址: {API_ENDPOINT}")
    print(f"数据条数: {len(posts)}")
    
    confirm = input("\n确认推送数据到数据库? (y/N): ").strip().lower()
    if confirm != 'y':
        print("取消操作")
        return
    
    push_to_api(posts)

if __name__ == "__main__":
    main()