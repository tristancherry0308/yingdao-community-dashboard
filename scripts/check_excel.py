import openpyxl

wb = openpyxl.load_workbook('../影刀社区帖子数据step03.xlsx')
ws = wb.active

headers = [cell.value for cell in ws[1]]
print("Excel Headers:", headers)
print("Row count:", ws.max_row)
print("Column count:", ws.max_column)

for row in ws.iter_rows(min_row=2, max_row=3, values_only=True):
    print("Sample row:", row)